#!/usr/bin/env python3
"""
GPT Monthly Usage Report Generator

Usage:
  python gpt_monthly_report.py input.xlsx --out GPT_Monthly_Report.xlsx

Assumptions:
  - Input workbook contains a sheet named "GPTData" with these columns:
      cadence, period_start, period_end, gpt_name, config_type, gpt_description, is_active,
      first_day_active_in_period, last_day_active_in_period, messages_workspace,
      unique_messagers_workspace, gpt_creator_email
  - The report excludes any rows whose GPT name or description matches:
      - smart[ -]?pack                 (Smart Pack GPTs)
      - \\bpsych                       (psychometric / psychology-related assessments)
      - exact GPT name: Candidate Letter Generator
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule


DEFAULT_SMART_RE = re.compile(r"smart[\s\-]*pack", re.I)
DEFAULT_PSYCH_RE = re.compile(r"\bpsych", re.I)
CANDIDATE_LETTER_NAME = "candidate letter generator"

# Color scheme
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, color="000000", size=10)
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")
POSITIVE_FILL = PatternFill("solid", fgColor="C6EFCE")
NEGATIVE_FILL = PatternFill("solid", fgColor="FFC7CE")
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def load_gptdata(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name="GPTData", engine="openpyxl")


def split_clean_excluded(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = df.copy()
    df["gpt_name"] = df.get("gpt_name", "").fillna("").astype(str)
    df["gpt_description"] = df.get("gpt_description", "").fillna("").astype(str)

    combined = (df["gpt_name"] + " " + df["gpt_description"]).astype(str)

    ex_smart = combined.str.contains(DEFAULT_SMART_RE)
    ex_psych = combined.str.contains(DEFAULT_PSYCH_RE)
    ex_candidate = df["gpt_name"].str.strip().str.lower().eq(CANDIDATE_LETTER_NAME)

    exclude = ex_smart | ex_psych | ex_candidate

    df["excluded"] = exclude
    df["exclude_reason"] = ""
    df.loc[ex_smart, "exclude_reason"] = "Smart Pack"
    df.loc[~ex_smart & ex_psych, "exclude_reason"] = "Psychometric"
    df.loc[~ex_smart & ~ex_psych & ex_candidate, "exclude_reason"] = "Candidate Letter Generator"

    df_clean = df.loc[~exclude].copy()
    df_excl = df.loc[exclude].copy()

    df_clean["period_start"] = pd.to_datetime(df_clean["period_start"], errors="coerce")
    df_excl["period_start"] = pd.to_datetime(df_excl["period_start"], errors="coerce")

    df_clean["period_key"] = df_clean["period_start"].dt.strftime("%Y-%m")
    df_excl["period_key"] = df_excl["period_start"].dt.strftime("%Y-%m")

    for col in ["messages_workspace", "unique_messagers_workspace", "is_active"]:
        if col in df_clean.columns:
            df_clean[col] = pd.to_numeric(df_clean[col], errors="coerce").fillna(0).astype(int)
        if col in df_excl.columns:
            df_excl[col] = pd.to_numeric(df_excl[col], errors="coerce").fillna(0).astype(int)

    return df_clean, df_excl


def _autosize(ws, max_width: int = 60, min_width: int = 8) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min(max_len + 2, max_width), min_width)


def _style_header_row(ws, row_num: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN


def _style_subheader(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER_THIN


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, 
              format_numbers: dict | None = None, add_totals: bool = False) -> int:
    """Write DataFrame to worksheet with formatting. Returns next available row."""
    
    # Headers
    for j, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E0E0E0")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    
    # Data rows
    for i, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(vertical="center")
            
            # Apply number formatting
            if format_numbers and df.columns[j - start_col] in format_numbers:
                cell.number_format = format_numbers[df.columns[j - start_col]]
    
    # Add totals row if requested
    next_row = start_row + len(df) + 1
    if add_totals and len(df) > 0:
        total_row = start_row + len(df) + 1
        ws.cell(row=total_row, column=start_col, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=start_col).border = BORDER_THIN
        
        for j, col_name in enumerate(df.columns[1:], start=start_col + 1):  # Skip first column
            if pd.api.types.is_numeric_dtype(df[col_name]):
                total = df[col_name].sum()
                cell = ws.cell(row=total_row, column=j, value=total)
                cell.font = Font(bold=True)
                cell.border = BORDER_THIN
                if format_numbers and col_name in format_numbers:
                    cell.number_format = format_numbers[col_name]
        next_row = total_row + 1
    
    ws.freeze_panes = f"{get_column_letter(start_col)}{start_row + 1}"
    return next_row


def calculate_trend_metrics(df_clean: pd.DataFrame, latest_period: str) -> dict:
    """Calculate month-over-month trends."""
    periods = sorted(df_clean["period_key"].dropna().unique())
    
    if len(periods) < 2:
        return {"has_trend": False}
    
    prev_period = periods[-2] if latest_period == periods[-1] else None
    if not prev_period:
        return {"has_trend": False}
    
    curr_data = df_clean[df_clean["period_key"] == latest_period]
    prev_data = df_clean[df_clean["period_key"] == prev_period]
    
    # Live config trends
    curr_live = curr_data[curr_data["config_type"].astype(str).str.lower() == "live"]
    prev_live = prev_data[prev_data["config_type"].astype(str).str.lower() == "live"]
    
    metrics = {
        "has_trend": True,
        "prev_period": prev_period,
        "messages_change": curr_live["messages_workspace"].sum() - prev_live["messages_workspace"].sum(),
        "messages_pct_change": (curr_live["messages_workspace"].sum() / prev_live["messages_workspace"].sum() - 1) 
                              if prev_live["messages_workspace"].sum() > 0 else 0,
        "active_gpts_change": (curr_live["messages_workspace"] > 0).sum() - (prev_live["messages_workspace"] > 0).sum(),
        "total_gpts_change": len(curr_live) - len(prev_live),
        "new_gpts": len(set(curr_live["gpt_name"]) - set(prev_live["gpt_name"])),
        "dropped_gpts": len(set(prev_live["gpt_name"]) - set(curr_live["gpt_name"])),
    }
    
    return metrics


def get_usage_tiers(live: pd.DataFrame) -> pd.DataFrame:
    """Categorize GPTs into usage tiers."""
    def categorize(messages):
        if messages == 0:
            return "Zero Usage"
        elif messages < 10:
            return "Low (1-9)"
        elif messages < 100:
            return "Medium (10-99)"
        elif messages < 1000:
            return "High (100-999)"
        else:
            return "Very High (1000+)"
    
    live = live.copy()
    live["usage_tier"] = live["messages_workspace"].apply(categorize)
    
    tier_summary = live.groupby("usage_tier").agg(
        gpt_count=("gpt_name", "nunique"),
        total_messages=("messages_workspace", "sum"),
        avg_messages=("messages_workspace", "mean"),
        total_users=("unique_messagers_workspace", "sum")
    ).reset_index()
    
    # Order tiers logically
    tier_order = ["Zero Usage", "Low (1-9)", "Medium (10-99)", "High (100-999)", "Very High (1000+)"]
    tier_summary["sort_key"] = tier_summary["usage_tier"].map({t: i for i, t in enumerate(tier_order)})
    tier_summary = tier_summary.sort_values("sort_key").drop("sort_key", axis=1)
    
    return tier_summary


def get_creator_segments(live: pd.DataFrame) -> pd.DataFrame:
    """Analyze creator segments."""
    creator_stats = live.groupby("gpt_creator_email").agg(
        gpts_created=("gpt_name", "nunique"),
        total_messages=("messages_workspace", "sum"),
        total_users=("unique_messagers_workspace", "sum"),
        avg_messages_per_gpt=("messages_workspace", "mean"),
        max_messages_single_gpt=("messages_workspace", "max")
    ).reset_index()
    
    def segment(row):
        if row["gpts_created"] == 1:
            return "Single GPT"
        elif row["gpts_created"] <= 3:
            return "Small Portfolio (2-3)"
        elif row["gpts_created"] <= 10:
            return "Medium Portfolio (4-10)"
        else:
            return "Power Creator (10+)"
    
    creator_stats["segment"] = creator_stats.apply(segment, axis=1)
    
    segment_summary = creator_stats.groupby("segment").agg(
        creators=("gpt_creator_email", "count"),
        total_gpts=("gpts_created", "sum"),
        total_messages=("total_messages", "sum"),
        avg_messages_per_creator=("total_messages", "mean")
    ).reset_index()
    
    # Order segments
    seg_order = ["Single GPT", "Small Portfolio (2-3)", "Medium Portfolio (4-10)", "Power Creator (10+)"]
    segment_summary["sort_key"] = segment_summary["segment"].map({s: i for i, s in enumerate(seg_order)})
    segment_summary = segment_summary.sort_values("sort_key").drop("sort_key", axis=1)
    
    return segment_summary, creator_stats


def build_report(input_path: Path, out_path: Path) -> None:
    df = load_gptdata(input_path)
    df_clean, df_excl = split_clean_excluded(df)

    # Get periods
    periods = sorted(df_clean["period_key"].dropna().unique())
    if not periods:
        raise ValueError("No valid periods found in data")
    
    latest_period = periods[-1]
    prev_period = periods[-2] if len(periods) > 1 else None
    
    # Calculate trends
    trend_metrics = calculate_trend_metrics(df_clean, latest_period)
    
    # Split live/draft
    live = df_clean[df_clean["config_type"].astype(str).str.lower() == "live"].copy()
    draft = df_clean[df_clean["config_type"].astype(str).str.lower() == "draft"].copy()
    
    # Core metrics
    tot_live_messages = int(live["messages_workspace"].sum())
    tot_all_messages = int(df_clean["messages_workspace"].sum())
    tot_draft_messages = int(draft["messages_workspace"].sum())
    
    gpts_with_usage_live = int((live["messages_workspace"] > 0).sum())
    gpts_with_usage_all = int((df_clean["messages_workspace"] > 0).sum())
    total_live_gpts = len(live)
    
    # Concentration metrics
    top5_live = live.sort_values("messages_workspace", ascending=False).head(5)
    top5_share = (top5_live["messages_workspace"].sum() / tot_live_messages) if tot_live_messages else 0.0
    
    top10_live = live.sort_values("messages_workspace", ascending=False).head(10)
    top10_share = (top10_live["messages_workspace"].sum() / tot_live_messages) if tot_live_messages else 0.0
    
    draft_share = float(tot_draft_messages / tot_all_messages) if tot_all_messages else 0.0
    
    # Engagement metrics
    live["messages_per_user"] = (live["messages_workspace"] / live["unique_messagers_workspace"].replace(0, pd.NA)).fillna(0)
    avg_engagement = live[live["messages_workspace"] > 0]["messages_per_user"].mean()
    
    # Usage tiers
    usage_tiers = get_usage_tiers(live)
    
    # Creator analysis
    creator_segments, creator_detail = get_creator_segments(live)
    
    # Top performers
    top10 = (
        live.sort_values("messages_workspace", ascending=False)
        .head(10)[["gpt_name", "messages_workspace", "unique_messagers_workspace", 
                   "messages_per_user", "gpt_creator_email"]]
        .rename(
            columns={
                "messages_workspace": "messages",
                "unique_messagers_workspace": "unique_users",
                "messages_per_user": "msgs_per_user",
                "gpt_creator_email": "creator_email",
            }
        )
    )
    
    top10["share_of_total"] = top10["messages"] / tot_live_messages
    
    # Top creators
    creator_summary = (
        live.groupby("gpt_creator_email")
        .agg(gpts=("gpt_name", "nunique"), 
             messages=("messages_workspace", "sum"),
             users=("unique_messagers_workspace", "sum"))
        .reset_index()
        .rename(columns={"gpt_creator_email": "creator_email"})
        .sort_values("messages", ascending=False)
        .head(10)
    )
    
    # Rising stars (new GPTs with high usage)
    if trend_metrics["has_trend"]:
        prev_live = df_clean[(df_clean["period_key"] == trend_metrics["prev_period"]) & 
                            (df_clean["config_type"].astype(str).str.lower() == "live")]
        prev_gpts = set(prev_live["gpt_name"])
        new_gpts = live[~live["gpt_name"].isin(prev_gpts)]
        rising_stars = new_gpts.nlargest(5, "messages_workspace")[["gpt_name", "messages_workspace", "gpt_creator_email"]]
    else:
        rising_stars = pd.DataFrame()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # ===== DASHBOARD SHEET =====
    dash = wb.create_sheet("Dashboard")
    
    # Title
    dash["A1"] = "GPT Usage Monthly Report"
    dash["A1"].font = Font(bold=True, size=16, color="1F4E78")
    dash["A2"] = f"Excludes: Smart Pack | Psychometric | Candidate Letter Generator | Generated: {datetime.now().strftime('%Y-%m-%d')}"
    dash["A2"].font = Font(italic=True, size=9, color="666666")
    
    # Period info
    dash["A4"] = "Reporting Period"
    dash["A4"].font = Font(bold=True, size=11)
    dash["B4"] = latest_period
    
    if trend_metrics["has_trend"]:
        dash["C4"] = f"vs {trend_metrics['prev_period']}"
        dash["C4"].font = Font(italic=True, color="666666")
    
    # KPI Section
    row = 6
    _style_subheader(dash, row, 1, "KEY PERFORMANCE INDICATORS")
    dash.merge_cells(f"A{row}:E{row}")
    
    # KPI Headers
    row += 1
    headers = ["Metric", "Live Only", "Trend", "Live + Draft", "Trend"]
    for j, h in enumerate(headers, 1):
        cell = dash.cell(row=row, column=j, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E0E0E0")
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER_THIN
    
    # KPI Data
    metrics_data = [
        ("Total Messages", tot_live_messages, tot_all_messages),
        ("GPTs with Usage", gpts_with_usage_live, gpts_with_usage_all),
        ("Total GPT Configs", total_live_gpts, len(df_clean)),
        ("Avg Messages per User", round(avg_engagement, 1), "-"),
    ]
    
    for metric, live_val, all_val in metrics_data:
        row += 1
        dash.cell(row=row, column=1, value=metric).border = BORDER_THIN
        dash.cell(row=row, column=2, value=live_val).border = BORDER_THIN
        dash.cell(row=row, column=2).number_format = "#,##0" if isinstance(live_val, (int, float)) else "0.0"
        dash.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        
        # Trend arrows for live
        if trend_metrics["has_trend"] and metric == "Total Messages":
            change = trend_metrics["messages_change"]
            pct = trend_metrics["messages_pct_change"]
            symbol = "▲" if change > 0 else "▼" if change < 0 else "→"
            trend_text = f"{symbol} {abs(pct):.1%}"
            dash.cell(row=row, column=3, value=trend_text).border = BORDER_THIN
            dash.cell(row=row, column=3).font = Font(color="008000" if change > 0 else "FF0000")
        
        dash.cell(row=row, column=4, value=all_val).border = BORDER_THIN
        dash.cell(row=row, column=4).number_format = "#,##0" if isinstance(all_val, (int, float)) else "@"
        dash.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    
    # Concentration metrics
    row += 2
    _style_subheader(dash, row, 1, "CONCENTRATION ANALYSIS")
    dash.merge_cells(f"A{row}:E{row}")
    
    row += 1
    dash.cell(row=row, column=1, value="Top 5 GPTs Share").border = BORDER_THIN
    dash.cell(row=row, column=2, value=top5_share).border = BORDER_THIN
    dash.cell(row=row, column=2).number_format = "0.0%"
    dash.cell(row=row, column=3, value="High" if top5_share > 0.5 else "Moderate" if top5_share > 0.3 else "Distributed")
    dash.cell(row=row, column=3).border = BORDER_THIN
    
    row += 1
    dash.cell(row=row, column=1, value="Top 10 GPTs Share").border = BORDER_THIN
    dash.cell(row=row, column=2, value=top10_share).border = BORDER_THIN
    dash.cell(row=row, column=2).number_format = "0.0%"
    
    row += 1
    dash.cell(row=row, column=1, value="Draft Share of Total").border = BORDER_THIN
    dash.cell(row=row, column=2, value=draft_share).border = BORDER_THIN
    dash.cell(row=row, column=2).number_format = "0.0%"
    
    # Top 10 Table
    row += 3
    _style_subheader(dash, row, 1, f"TOP 10 GPTs BY MESSAGES ({latest_period})")
    dash.merge_cells(f"A{row}:F{row}")
    
    row += 1
    _write_df(dash, top10, start_row=row, start_col=1,
              format_numbers={"messages": "#,##0", "unique_users": "#,##0", 
                            "msgs_per_user": "0.0", "share_of_total": "0.0%"})
    
    _autosize(dash)
    
    # ===== INSIGHTS SHEET =====
    ins = wb.create_sheet("Insights")
    
    ins["A1"] = "Strategic Insights & Analysis"
    ins["A1"].font = Font(bold=True, size=14, color="1F4E78")
    
    ins["A3"] = f"Analysis Period: {latest_period}"
    ins["A3"].font = Font(bold=True)
    
    current_row = 5
    
    # Executive Summary
    _style_subheader(ins, current_row, 1, "EXECUTIVE SUMMARY")
    ins.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 1
    
    insights_text = []
    
    # Growth insight
    if trend_metrics["has_trend"]:
        growth = trend_metrics["messages_pct_change"]
        if abs(growth) > 0.1:
            direction = "surged" if growth > 0 else "declined"
            insights_text.append(f"• Message volume has {direction} by {abs(growth):.1%} compared to {trend_metrics['prev_period']}")
        else:
            insights_text.append(f"• Message volume remained stable compared to {trend_metrics['prev_period']}")
    
    # Concentration insight
    if top5_share > 0.6:
        insights_text.append(f"• High concentration risk: Top 5 GPTs account for {top5_share:.1%} of all usage")
    elif top5_share < 0.3:
        insights_text.append(f"• Healthy distribution: Top 5 GPTs account for only {top5_share:.1%} of usage")
    
    # Draft insight
    if draft_share > 0.2:
        insights_text.append(f"• Significant draft activity: {draft_share:.1%} of messages are from draft configs")
    
    # Zero usage insight
    zero_pct = (live["messages_workspace"] == 0).sum() / len(live)
    if zero_pct > 0.3:
        insights_text.append(f"• Cleanup opportunity: {zero_pct:.1%} of live GPTs have zero usage")
    
    for text in insights_text:
        ins.cell(row=current_row, column=1, value=text)
        ins.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        current_row += 1
    
    current_row += 2
    
    # Usage Tiers Table
    _style_subheader(ins, current_row, 1, "USAGE DISTRIBUTION (LIVE GPTs)")
    ins.merge_cells(f"A{current_row}:E{current_row}")
    current_row += 1
    
    next_row = _write_df(ins, usage_tiers, start_row=current_row, start_col=1,
                        format_numbers={"gpt_count": "#,##0", "total_messages": "#,##0",
                                      "avg_messages": "#,##0.0", "total_users": "#,##0"},
                        add_totals=True)
    current_row = next_row + 1
    
    # Creator Segments Table
    _style_subheader(ins, current_row, 1, "CREATOR ECOSYSTEM ANALYSIS")
    ins.merge_cells(f"A{current_row}:E{current_row}")
    current_row += 1
    
    next_row = _write_df(ins, creator_segments, start_row=current_row, start_col=1,
                        format_numbers={"creators": "#,##0", "total_gpts": "#,##0",
                                      "total_messages": "#,##0", "avg_messages_per_creator": "#,##0.0"},
                        add_totals=True)
    current_row = next_row + 1
    
    # Top Creators Table
    _style_subheader(ins, current_row, 1, "TOP 10 CREATORS BY MESSAGE VOLUME")
    ins.merge_cells(f"A{current_row}:E{current_row}")
    current_row += 1
    
    _write_df(ins, creator_summary, start_row=current_row, start_col=1,
              format_numbers={"gpts": "#,##0", "messages": "#,##0", "users": "#,##0"})
    current_row += 12
    
    # Rising Stars
    if not rising_stars.empty:
        _style_subheader(ins, current_row, 1, f"RISING STARS (New GPTs in {latest_period})")
        ins.merge_cells(f"A{current_row}:D{current_row}")
        current_row += 1
        
        rising_formatted = rising_stars.rename(columns={
            "gpt_name": "GPT Name",
            "messages_workspace": "Messages",
            "gpt_creator_email": "Creator"
        })
        _write_df(ins, rising_formatted, start_row=current_row, start_col=1,
                  format_numbers={"Messages": "#,##0"})
    
    _autosize(ins)
    
    # ===== TRENDS SHEET (if multi-month) =====
    if len(periods) > 1:
        trends = wb.create_sheet("Trends")
        
        # Monthly trend data
        monthly_summary = df_clean.groupby(["period_key", "config_type"]).agg(
            total_messages=("messages_workspace", "sum"),
            active_gpts=("gpt_name", lambda x: (df_clean.loc[x.index, "messages_workspace"] > 0).sum()),
            total_gpts=("gpt_name", "nunique")
        ).reset_index()
        
        # Pivot for easier reading
        pivot_msgs = monthly_summary.pivot(index="period_key", columns="config_type", 
                                          values="total_messages").fillna(0)
        pivot_msgs["Total"] = pivot_msgs.sum(axis=1)
        pivot_msgs = pivot_msgs.reset_index()
        
        trends["A1"] = "Monthly Message Trends"
        trends["A1"].font = Font(bold=True, size=12)
        _write_df(trends, pivot_msgs, start_row=2, start_col=1,
                  format_numbers={"live": "#,##0", "draft": "#,##0", "Total": "#,##0"})
        
        _autosize(trends)
    
    # ===== DATA SHEETS =====
    dc = wb.create_sheet("Data_Clean")
    _write_df(dc, df_clean, format_numbers={"messages_workspace": "#,##0", 
                                           "unique_messagers_workspace": "#,##0"})
    _autosize(dc)
    
    ex = wb.create_sheet("Excluded_Rows")
    _write_df(ex, df_excl)
    _autosize(ex)
    
    # ===== DETAILED CREATORS SHEET =====
    creators = wb.create_sheet("Creator_Detail")
    _write_df(creators, creator_detail.sort_values("total_messages", ascending=False),
              format_numbers={"gpts_created": "#,##0", "total_messages": "#,##0",
                            "total_users": "#,##0", "avg_messages_per_gpt": "#,##0.0",
                            "max_messages_single_gpt": "#,##0"})
    _autosize(creators)
    
    wb.save(out_path)
    print(f"✓ Report generated: {out_path}")
    print(f"  - Period: {latest_period}")
    print(f"  - Live GPTs: {total_live_gpts:,} ({gpts_with_usage_live:,} with usage)")
    print(f"  - Total messages: {tot_live_messages:,}")
    if trend_metrics["has_trend"]:
        print(f"  - MoM change: {trend_metrics['messages_pct_change']:+.1%}")


def main() -> None:
    p = argparse.ArgumentParser(description="Generate GPT Monthly Usage Report")
    p.add_argument("input", type=Path, help="Path to input Excel export (.xlsx) containing GPTData sheet")
    p.add_argument("--out", type=Path, default=Path("GPT_Monthly_Report.xlsx"), 
                   help="Output report .xlsx path")
    args = p.parse_args()

    if not args.input.exists():
        print(f"Error: Input file not found: {args.input}")
        return
    
    build_report(args.input, args.out)


if __name__ == "__main__":
    main()